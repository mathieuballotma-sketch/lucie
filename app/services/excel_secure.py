"""
ExcelSecureLoader v2 — Production-ready secure Excel file loader.

Security layers (in order):
  1. Extension check — block macro-capable extensions (.xlsm, .xls, etc.)
  2. ZIP inspection — vbaProject.bin detection in OOXML packages
  3. OLE compound file analysis — VBA storage in .xlsb / .xls
  4. Formula injection scanning — DDE, dangerous functions, injection via leading chars
  5. Array formula detection — {=...} patterns
  6. Safe text export — to_llm_text() guarantees no executable formula
  7. EventBus integration — _publish_threat() on detection

Usage:
    loader = ExcelSecureLoader()
    rows, report = loader.load("invoice.xlsx")
    safe_text = loader.to_llm_text(rows)
"""
from __future__ import annotations

import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Security constants
# ---------------------------------------------------------------------------

DANGEROUS_FUNCTIONS: frozenset[str] = frozenset({
    "HYPERLINK", "WEBSERVICE", "FILTERXML", "RTD", "CALL", "REGISTER",
    "EXEC", "SHELL", "DDE", "DDEAUTO", "INDIRECT", "INFO", "REGISTER.ID",
})

DDE_PATTERN = re.compile(
    r"\bDDE(?:AUTO)?\s*\(",
    re.IGNORECASE,
)

FUNC_PATTERN = re.compile(
    r"=\s*(?:" + "|".join(re.escape(f) for f in sorted(DANGEROUS_FUNCTIONS)) + r")\s*\(",
    re.IGNORECASE,
)

# Leading-character injection: =, +, -, @, TAB, CR
INJECTION_VIA_FORMULA = re.compile(r"^[=+\-@\t\r]")

# Array formula: {=...}
ARRAY_FORMULA_PATTERN = re.compile(r"^\{=.*\}$")

_MACRO_EXTENSIONS: frozenset[str] = frozenset({
    ".xlsm", ".xltm", ".xlam", ".xls", ".xlt", ".xla",
})
_XLSB_EXTENSION = ".xlsb"


# ---------------------------------------------------------------------------
# Threat report
# ---------------------------------------------------------------------------

@dataclass
class ThreatReport:
    filename: str
    threats: list[str] = field(default_factory=list)
    has_macros: bool = False
    has_dangerous_formula: bool = False
    blocked_cells: int = 0

    @property
    def is_safe(self) -> bool:
        return not self.threats

    def to_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "threats": self.threats,
            "has_macros": self.has_macros,
            "has_dangerous_formula": self.has_dangerous_formula,
            "blocked_cells": self.blocked_cells,
        }


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

class ExcelSecureLoader:
    """
    Loads .xlsx and .xlsb Excel files with multi-layer security checks.

    - Never executes formulas (data_only=True for openpyxl)
    - Rejects files with detected macros (raises ValueError)
    - to_llm_text() replaces all formula-like cells with [FORMULA_BLOCKED]
    """

    def __init__(self, event_bus: Any | None = None) -> None:
        self._event_bus = event_bus

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def load(self, path: str | Path) -> tuple[list[list[Any]], ThreatReport]:
        """
        Load an Excel file and return (rows, ThreatReport).

        Raises:
            ValueError: if macros are detected.
            RuntimeError: if a required library is missing.
        """
        path = Path(path)
        report = ThreatReport(filename=path.name)

        self._check_extension(path, report)
        data = self._load_data(path, report)
        self._scan_formulas(data, report)

        if report.has_macros:
            self._publish_threat(report)
            raise ValueError(
                f"Macros detected in '{path.name}' — file rejected for security. "
                f"Threats: {report.threats}"
            )

        if report.threats:
            self._publish_threat(report)

        return data, report

    def to_llm_text(
        self,
        rows: list[list[Any]],
        sheet_name: str = "Sheet",
        max_rows: int = 1000,
    ) -> str:
        """
        Convert rows to plain-text safe for LLM consumption.

        All formula-like cells are replaced with [FORMULA_BLOCKED].
        No formula is ever passed as-is to the output.
        """
        lines: list[str] = [f"# Sheet: {sheet_name}"]
        for row in rows[:max_rows]:
            sanitized: list[str] = []
            for cell in row:
                cell_str = "" if cell is None else str(cell)
                if self._is_dangerous_formula(cell_str):
                    sanitized.append("[FORMULA_BLOCKED]")
                else:
                    sanitized.append(cell_str)
            lines.append(" | ".join(sanitized))
        if len(rows) > max_rows:
            lines.append(f"[...{len(rows) - max_rows} more rows truncated]")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Extension check
    # ------------------------------------------------------------------

    def _check_extension(self, path: Path, report: ThreatReport) -> None:
        ext = path.suffix.lower()
        if ext in _MACRO_EXTENSIONS:
            report.has_macros = True
            report.threats.append(f"Macro-capable file extension: {ext}")

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    def _load_data(self, path: Path, report: ThreatReport) -> list[list[Any]]:
        ext = path.suffix.lower()
        if ext == _XLSB_EXTENSION:
            return self._load_xlsb(path, report)
        return self._load_xlsx(path, report)

    def _load_xlsx(self, path: Path, report: ThreatReport) -> list[list[Any]]:
        """Load .xlsx (or OOXML variants) with openpyxl in read_only + data_only mode."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required: pip install openpyxl")

        self._inspect_zip_for_macros(path, report)

        wb = openpyxl.load_workbook(
            filename=str(path),
            read_only=True,
            keep_vba=False,
            data_only=True,
        )
        rows: list[list[Any]] = []
        ws = wb.active
        if ws is None and wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        if ws is not None:
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
        wb.close()
        return rows

    def _load_xlsb(self, path: Path, report: ThreatReport) -> list[list[Any]]:
        """Load .xlsb with pyxlsb."""
        try:
            import pyxlsb
        except ImportError:
            raise RuntimeError(
                "pyxlsb is required for .xlsb files: pip install pyxlsb"
            )

        self._inspect_ole_for_macros(path, report)

        rows: list[list[Any]] = []
        with pyxlsb.open_workbook(str(path)) as wb:
            if wb.sheets:
                with wb.get_sheet(1) as ws:
                    for row in ws.rows():
                        rows.append([cell.v for cell in row])
        return rows

    # ------------------------------------------------------------------
    # Macro detection
    # ------------------------------------------------------------------

    def _inspect_zip_for_macros(self, path: Path, report: ThreatReport) -> None:
        """Check the OOXML ZIP container for vbaProject.bin."""
        try:
            with zipfile.ZipFile(path, "r") as zf:
                for name in zf.namelist():
                    if "vbaProject.bin" in name or name.lower().endswith("/vba/"):
                        report.has_macros = True
                        report.threats.append(
                            f"vbaProject.bin found in ZIP archive: {name}"
                        )
        except zipfile.BadZipFile:
            pass
        except Exception as exc:
            logger.warning("ZIP macro inspection failed for %s: %s", path.name, exc)

    def _inspect_ole_for_macros(self, path: Path, report: ThreatReport) -> None:
        """Check xlsb/xls OLE compound file for VBA storage entries."""
        try:
            import olefile
        except ImportError:
            logger.debug("olefile not available — OLE macro check skipped")
            return

        try:
            if not olefile.isOleFile(str(path)):
                return
            with olefile.OleFileIO(str(path)) as ole:
                for entry in ole.listdir():
                    entry_str = "/".join(entry).lower()
                    if "vba" in entry_str or "macros" in entry_str:
                        report.has_macros = True
                        report.threats.append(
                            f"OLE VBA storage entry found: {'/'.join(entry)}"
                        )
        except Exception as exc:
            logger.warning("OLE inspection failed for %s: %s", path.name, exc)

    # ------------------------------------------------------------------
    # Formula scanning
    # ------------------------------------------------------------------

    def _scan_formulas(
        self, rows: list[list[Any]], report: ThreatReport
    ) -> None:
        """Scan all cells for dangerous formulas and injection patterns."""
        for row in rows:
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell)
                if self._is_dangerous_formula(cell_str):
                    report.has_dangerous_formula = True
                    report.blocked_cells += 1
                    report.threats.append(
                        f"Dangerous formula/injection: {cell_str[:80]}"
                    )

    def _is_dangerous_formula(self, value: str) -> bool:
        if not value:
            return False
        if ARRAY_FORMULA_PATTERN.match(value):
            return True
        if DDE_PATTERN.search(value):
            return True
        if FUNC_PATTERN.search(value):
            return True
        if INJECTION_VIA_FORMULA.match(value):
            return True
        return False

    # ------------------------------------------------------------------
    # EventBus integration
    # ------------------------------------------------------------------

    def _publish_threat(self, report: ThreatReport) -> None:
        """Publish a threat detection event to the EventBus if configured."""
        if self._event_bus is None:
            return
        try:
            import asyncio
            payload = report.to_dict()
            loop = None
            try:
                loop = asyncio.get_running_loop()
            except RuntimeError:
                pass

            if loop and loop.is_running():
                asyncio.ensure_future(
                    self._event_bus.publish(
                        channel="security.excel.threat",
                        data=payload,
                        source="ExcelSecureLoader",
                    )
                )
            else:
                asyncio.run(
                    self._event_bus.publish(
                        channel="security.excel.threat",
                        data=payload,
                        source="ExcelSecureLoader",
                    )
                )
        except Exception as exc:
            logger.warning("EventBus publish failed: %s", exc)
