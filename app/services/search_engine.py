"""
Moteur de recherche local intelligent pour Agent Lucide.
Indexe les fichiers du Mac, genere des mots-cles via LLM local,
et permet une recherche semantique via FAISS + FTS5.

Composants :
  - DocumentIndex : structure d'un document indexe
  - FileExtractor : extraction de contenu par type de fichier
  - KeywordGenerator : generation de mots-cles via LLM local
  - SearchEngine : moteur principal (FAISS + SQLite FTS5)
"""

import asyncio
import hashlib
import json
import sqlite3
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import faiss
import numpy as np

from ..utils.json_parser import safe_json_loads
from ..utils.logger import get_logger

logger = get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DocumentIndex
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class DocumentIndex:
    """Representation d'un document indexe."""

    file_path: str
    file_name: str
    file_type: str
    file_size: int
    modified_at: float
    indexed_at: float
    content_hash: str
    keywords: List[str] = field(default_factory=list)
    summary: str = ""
    category: str = "other"
    content_preview: str = ""
    embedding_id: Optional[int] = None


# ─────────────────────────────────────────────────────────────────────────────
# FileExtractor
# ─────────────────────────────────────────────────────────────────────────────
class FileExtractor:
    """Extrait le contenu textuel de differents types de fichiers."""

    SUPPORTED_EXTENSIONS: set[str] = {
        ".txt", ".md", ".py", ".js", ".ts", ".html", ".css", ".json", ".yaml", ".yml",
        ".toml", ".cfg", ".ini", ".sh", ".bash", ".zsh",
        ".csv", ".log", ".xml", ".sql",
        ".pdf", ".docx", ".xlsx",
        ".rst", ".tex",
    }

    def __init__(self, max_file_size: int = 10_000_000) -> None:
        self._max_file_size = max_file_size

    async def extract(self, file_path: str) -> Optional[str]:
        """Extrait le texte d'un fichier. Retourne None si non supporte."""
        p = Path(file_path)
        if not p.exists() or not p.is_file():
            return None

        ext = p.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            return None

        # Verifier la taille
        try:
            if p.stat().st_size > self._max_file_size:
                logger.debug(f"Fichier trop gros, skip: {file_path}")
                return None
        except OSError:
            return None

        if ext == ".pdf":
            return await self._extract_pdf(file_path)
        elif ext == ".docx":
            return await self._extract_docx(file_path)
        elif ext == ".xlsx":
            return await self._extract_xlsx(file_path)
        else:
            return await self._extract_text(file_path)

    async def _extract_text(self, file_path: str) -> Optional[str]:
        """Lit un fichier texte avec fallback encodage."""
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._read_text_sync, file_path)

    def _read_text_sync(self, file_path: str) -> Optional[str]:
        """Lecture synchrone avec multi-encodage."""
        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(file_path, "r", encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.debug(f"Erreur lecture {file_path}: {e}")
                return None
        return None

    async def _extract_pdf(self, file_path: str) -> Optional[str]:
        """Extrait le texte d'un PDF via pymupdf."""
        try:
            import pymupdf
            loop = asyncio.get_running_loop()

            def _read() -> str:
                doc: Any = pymupdf.open(file_path)  # type: ignore[no-untyped-call]
                text = ""
                for page in doc:
                    text += str(page.get_text())
                doc.close()
                return text

            return await loop.run_in_executor(None, _read)
        except ImportError:
            logger.debug("pymupdf non disponible, PDF ignore")
            return None
        except Exception as e:
            logger.warning(f"Erreur PDF {file_path}: {e}")
            return None

    async def _extract_docx(self, file_path: str) -> Optional[str]:
        """Extrait le texte d'un DOCX via python-docx."""
        try:
            import docx
            loop = asyncio.get_running_loop()

            def _read() -> str:
                doc = docx.Document(file_path)
                return "\n".join(p.text for p in doc.paragraphs)

            return await loop.run_in_executor(None, _read)
        except ImportError:
            logger.debug("python-docx non disponible, DOCX ignore")
            return None
        except Exception as e:
            logger.warning(f"Erreur DOCX {file_path}: {e}")
            return None

    async def _extract_xlsx(self, file_path: str) -> Optional[str]:
        """Extrait le texte d'un XLSX via openpyxl."""
        try:
            import openpyxl
            loop = asyncio.get_running_loop()

            def _read() -> str:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                parts: List[str] = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells:
                            parts.append("\t".join(cells))
                wb.close()
                return "\n".join(parts)

            return await loop.run_in_executor(None, _read)
        except ImportError:
            logger.debug("openpyxl non disponible, XLSX ignore")
            return None
        except Exception as e:
            logger.warning(f"Erreur XLSX {file_path}: {e}")
            return None


# ─────────────────────────────────────────────────────────────────────────────
# KeywordGenerator
# ─────────────────────────────────────────────────────────────────────────────
class KeywordGenerator:
    """Genere des mots-cles et resumes pour un document via le LLM local."""

    KEYWORD_PROMPT: str = (
        "Analyse ce document et retourne UNIQUEMENT un JSON avec ce format exact :\n"
        '{{\n'
        '    "keywords": ["mot1", "mot2", "mot3", "mot4", "mot5"],\n'
        '    "summary": "Resume en une phrase de ce document.",\n'
        '    "category": "code|document|data|config|media|other"\n'
        '}}\n\n'
        "Document ({file_type}, {file_name}) :\n"
        "---\n"
        "{content_preview}\n"
        "---\n\n"
        "Reponds UNIQUEMENT avec le JSON, rien d'autre."
    )

    def __init__(self, provider_manager: Any) -> None:
        self._provider = provider_manager

    async def generate(
        self, file_name: str, file_type: str, content: str
    ) -> Dict[str, Any]:
        """Genere mots-cles + resume. Retourne un dict avec keywords, summary, category."""
        content_preview = content[:2000]
        prompt = self.KEYWORD_PROMPT.format(
            file_type=file_type,
            file_name=file_name,
            content_preview=content_preview,
        )

        try:
            loop = asyncio.get_running_loop()
            response = await loop.run_in_executor(
                None,
                lambda: self._provider.generate(
                    prompt=prompt,
                    priority="speed",
                    max_tokens=256,
                    temperature=0.3,
                ),
            )

            result = safe_json_loads(response, default=None, expected_type=dict)
            if result is None:
                return self._fallback(file_name, content)

            # Valider la structure
            keywords = result.get("keywords", [])
            if not isinstance(keywords, list):
                keywords = []
            summary = str(result.get("summary", ""))
            category = str(result.get("category", "other"))
            if category not in ("code", "document", "data", "config", "media", "other"):
                category = "other"

            return {
                "keywords": [str(k) for k in keywords[:10]],
                "summary": summary[:500],
                "category": category,
            }
        except Exception as e:
            logger.debug(f"LLM keyword generation failed: {e}")
            return self._fallback(file_name, content)

    def _fallback(self, file_name: str, content: str) -> Dict[str, Any]:
        """Fallback sans LLM : mots-cles extraits du nom de fichier."""
        name_parts = Path(file_name).stem.replace("_", " ").replace("-", " ").split()
        return {
            "keywords": name_parts[:5],
            "summary": content[:200] if content else "",
            "category": "other",
        }


# ─────────────────────────────────────────────────────────────────────────────
# SearchEngine
# ─────────────────────────────────────────────────────────────────────────────
class LocalSearchEngine:
    """Moteur de recherche local intelligent avec indexation FAISS + SQLite FTS5."""

    def __init__(
        self,
        index_dir: str = "data/search_index",
        embedder: Optional[Any] = None,
        provider_manager: Optional[Any] = None,
        excluded_dirs: Optional[List[str]] = None,
        excluded_extensions: Optional[List[str]] = None,
        max_file_size: int = 10_000_000,
        generate_keywords: bool = True,
    ) -> None:
        self._index_dir = Path(index_dir)
        self._index_dir.mkdir(parents=True, exist_ok=True)

        self._embedder = embedder
        self._extractor = FileExtractor(max_file_size=max_file_size)
        self._keyword_gen: Optional[KeywordGenerator] = None
        if provider_manager and generate_keywords:
            self._keyword_gen = KeywordGenerator(provider_manager)

        self._excluded_dirs = set(excluded_dirs or [
            ".git", "node_modules", "__pycache__", ".venv", "venv",
            ".Trash", ".Spotlight-V100", ".fseventsd",
        ])
        self._excluded_extensions = set(excluded_extensions or [
            ".pyc", ".pyo", ".so", ".dylib", ".o", ".a",
        ])

        self._lock = threading.Lock()
        self._dimension: int = 1024

        if embedder is not None:
            self._dimension = embedder.get_sentence_embedding_dimension()

        self._init_db()
        self._init_faiss()

        logger.info(
            f"LocalSearchEngine initialise (dir={index_dir}, "
            f"dim={self._dimension}, keywords={generate_keywords})"
        )

    def _init_db(self) -> None:
        """Initialise SQLite avec FTS5."""
        db_path = self._index_dir / "search.db"
        self._conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._conn.execute("PRAGMA journal_mode=WAL")
        self._conn.execute("PRAGMA synchronous=NORMAL")

        self._conn.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT UNIQUE NOT NULL,
                file_name TEXT NOT NULL,
                file_type TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                modified_at REAL NOT NULL,
                indexed_at REAL NOT NULL,
                content_hash TEXT NOT NULL,
                keywords TEXT NOT NULL DEFAULT '[]',
                summary TEXT DEFAULT '',
                category TEXT DEFAULT 'other',
                content_preview TEXT DEFAULT '',
                embedding_id INTEGER
            )
        """)
        self._conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_doc_file_type ON documents(file_type)"
        )
        self._conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_doc_category ON documents(category)"
        )

        # FTS5 pour recherche full-text
        try:
            self._conn.execute("""
                CREATE VIRTUAL TABLE IF NOT EXISTS documents_fts USING fts5(
                    file_name, keywords, summary, content_preview,
                    content='documents',
                    content_rowid='id'
                )
            """)
        except sqlite3.OperationalError:
            logger.warning("FTS5 non disponible, recherche par mots-cles degradee")

        self._conn.commit()

    def _init_faiss(self) -> None:
        """Initialise ou charge l'index FAISS."""
        self._faiss_path = self._index_dir / "search_faiss.index"
        self._mapping_path = self._index_dir / "search_mapping.json"
        self._id_mapping: List[int] = []

        if self._faiss_path.exists():
            try:
                self._faiss_index: Any = faiss.read_index(str(self._faiss_path))
                if self._faiss_index.d != self._dimension:
                    logger.warning("Dimension FAISS incompatible, recreation")
                    self._faiss_index = faiss.IndexFlatL2(self._dimension)
                else:
                    logger.info(
                        f"Index FAISS recherche charge ({self._faiss_index.ntotal} vecteurs)"
                    )
            except Exception as e:
                logger.warning(f"Erreur chargement FAISS recherche: {e}")
                self._faiss_index = faiss.IndexFlatL2(self._dimension)
        else:
            self._faiss_index = faiss.IndexFlatL2(self._dimension)

        if self._mapping_path.exists():
            try:
                self._id_mapping = json.loads(self._mapping_path.read_text())
            except Exception:
                self._id_mapping = []

    def _save_faiss(self) -> None:
        """Sauvegarde l'index FAISS et le mapping."""
        faiss.write_index(self._faiss_index, str(self._faiss_path))
        self._mapping_path.write_text(json.dumps(self._id_mapping))

    def _compute_hash(self, file_path: str) -> str:
        """Calcule le hash blake2b d'un fichier."""
        hasher = hashlib.blake2b(digest_size=16)
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    hasher.update(chunk)
        except OSError:
            return ""
        return hasher.hexdigest()

    def _is_excluded(self, file_path: Path) -> bool:
        """Verifie si un fichier est exclu."""
        # Extension exclue
        if file_path.suffix.lower() in self._excluded_extensions:
            return True
        # Dossier exclu
        for part in file_path.parts:
            if part in self._excluded_dirs:
                return True
        return False

    # ─────────────────────────────────────────────────────────────────────
    # Indexation
    # ─────────────────────────────────────────────────────────────────────
    async def add_directory(self, dir_path: str, recursive: bool = True) -> int:
        """Ajoute un dossier a l'index. Retourne le nombre de fichiers indexes."""
        p = Path(dir_path).expanduser().resolve()
        if not p.is_dir():
            logger.warning(f"Dossier invalide: {dir_path}")
            return 0

        count = 0
        if recursive:
            files = list(p.rglob("*"))
        else:
            files = list(p.iterdir())

        for file_path in files:
            if not file_path.is_file():
                continue
            if self._is_excluded(file_path):
                continue
            if file_path.suffix.lower() not in FileExtractor.SUPPORTED_EXTENSIONS:
                continue

            try:
                doc = await self.index_file(str(file_path))
                if doc is not None:
                    count += 1
            except Exception as e:
                logger.debug(f"Erreur indexation {file_path}: {e}")

        logger.info(f"Dossier indexe: {dir_path} ({count} fichiers)")
        return count

    async def index_file(self, file_path: str) -> Optional[DocumentIndex]:
        """Indexe un seul fichier. Retourne le DocumentIndex cree ou None."""
        p = Path(file_path)
        if not p.exists() or not p.is_file():
            return None

        # Hash pour detecter les changements
        content_hash = self._compute_hash(file_path)
        if not content_hash:
            return None

        # Verifier si deja indexe avec le meme hash
        with self._lock:
            row = self._conn.execute(
                "SELECT id FROM documents WHERE file_path = ? AND content_hash = ?",
                (file_path, content_hash),
            ).fetchone()
            if row:
                return None  # deja indexe, pas change

        # Extraire le contenu
        content = await self._extractor.extract(file_path)
        if content is None or not content.strip():
            return None

        content_preview = content[:500]
        stat = p.stat()

        # Generer mots-cles via LLM
        keywords: List[str] = []
        summary = ""
        category = "other"

        keyword_gen = self._keyword_gen
        if keyword_gen is not None:
            try:
                kw_result = await keyword_gen.generate(
                    p.name, p.suffix.lower(), content
                )
                keywords = kw_result.get("keywords", [])
                summary = kw_result.get("summary", "")
                category = kw_result.get("category", "other")
            except Exception as e:
                logger.debug(f"Keyword gen failed for {p.name}: {e}")

        # Generer l'embedding
        embedding_id: Optional[int] = None
        embedder = self._embedder
        if embedder is not None:
            try:
                # Combiner nom + resume + preview pour un embedding riche
                embed_text = f"{p.name} {summary} {content_preview}"
                loop = asyncio.get_running_loop()
                emb = await loop.run_in_executor(
                    None, lambda: embedder.encode(embed_text)
                )
                emb_array = np.array([emb]).astype("float32")

                with self._lock:
                    self._faiss_index.add(emb_array)
                    faiss_pos = self._faiss_index.ntotal - 1
                    embedding_id = faiss_pos
            except Exception as e:
                logger.debug(f"Embedding failed for {p.name}: {e}")

        # Sauvegarder en SQLite
        now = time.time()
        keywords_json = json.dumps(keywords)

        with self._lock:
            # Supprimer l'ancien si existe
            self._conn.execute(
                "DELETE FROM documents WHERE file_path = ?", (file_path,)
            )
            cur = self._conn.cursor()
            cur.execute(
                """
                INSERT INTO documents
                    (file_path, file_name, file_type, file_size, modified_at,
                     indexed_at, content_hash, keywords, summary, category,
                     content_preview, embedding_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    file_path,
                    p.name,
                    p.suffix.lower(),
                    stat.st_size,
                    stat.st_mtime,
                    now,
                    content_hash,
                    keywords_json,
                    summary,
                    category,
                    content_preview,
                    embedding_id,
                ),
            )
            doc_id = cur.lastrowid or 0

            # Mettre a jour FTS5
            try:
                self._conn.execute(
                    """
                    INSERT INTO documents_fts(rowid, file_name, keywords, summary, content_preview)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (doc_id, p.name, " ".join(keywords), summary, content_preview),
                )
            except sqlite3.OperationalError:
                pass  # FTS5 non disponible

            self._conn.commit()

            if embedding_id is not None:
                self._id_mapping.append(doc_id)
                self._save_faiss()

        doc = DocumentIndex(
            file_path=file_path,
            file_name=p.name,
            file_type=p.suffix.lower(),
            file_size=stat.st_size,
            modified_at=stat.st_mtime,
            indexed_at=now,
            content_hash=content_hash,
            keywords=keywords,
            summary=summary,
            category=category,
            content_preview=content_preview,
            embedding_id=embedding_id,
        )
        logger.debug(f"Indexe: {p.name} ({len(keywords)} keywords)")
        return doc

    # ─────────────────────────────────────────────────────────────────────
    # Recherche
    # ─────────────────────────────────────────────────────────────────────
    async def search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Recherche hybride : semantique FAISS + mots-cles FTS5."""
        semantic_results = await self._search_semantic(query, top_k * 2)
        keyword_results = await self._search_keywords(query, top_k * 2)

        # Fusionner avec score hybride
        merged: Dict[str, Dict[str, Any]] = {}

        for r in semantic_results:
            fp = r["file_path"]
            merged[fp] = r
            merged[fp]["score"] = r.get("semantic_score", 0) * 0.7

        for r in keyword_results:
            fp = r["file_path"]
            if fp in merged:
                merged[fp]["score"] += r.get("keyword_score", 0) * 0.3
            else:
                merged[fp] = r
                merged[fp]["score"] = r.get("keyword_score", 0) * 0.3

        results = sorted(merged.values(), key=lambda x: x.get("score", 0), reverse=True)
        return results[:top_k]

    async def _search_semantic(self, query: str, top_k: int) -> List[Dict[str, Any]]:
        """Recherche semantique via FAISS."""
        embedder = self._embedder
        if embedder is None or self._faiss_index.ntotal == 0:
            return []

        try:
            loop = asyncio.get_running_loop()
            q_emb = await loop.run_in_executor(
                None, lambda: embedder.encode(query)
            )
            q_array = np.array([q_emb]).astype("float32")
        except Exception as e:
            logger.debug(f"Erreur embedding recherche: {e}")
            return []

        n = min(top_k, self._faiss_index.ntotal)
        distances, indices = self._faiss_index.search(q_array, n)

        results: List[Dict[str, Any]] = []
        with self._lock:
            for i, idx in enumerate(indices[0]):
                if idx == -1 or idx >= len(self._id_mapping):
                    continue
                doc_id = self._id_mapping[idx]
                distance = float(distances[0][i])
                similarity = 1.0 / (1.0 + distance)

                row = self._conn.execute(
                    """
                    SELECT file_path, file_name, file_type, file_size,
                           keywords, summary, category
                    FROM documents WHERE id = ?
                    """,
                    (doc_id,),
                ).fetchone()

                if row:
                    results.append({
                        "file_path": row[0],
                        "file_name": row[1],
                        "file_type": row[2],
                        "file_size": row[3],
                        "keywords": json.loads(row[4]),
                        "summary": row[5],
                        "category": row[6],
                        "semantic_score": similarity,
                        "score": similarity,
                    })

        return results

    async def _search_keywords(self, query: str, top_k: int) -> List[Dict[str, Any]]:
        """Recherche par mots-cles via FTS5."""
        results: List[Dict[str, Any]] = []
        with self._lock:
            try:
                rows = self._conn.execute(
                    """
                    SELECT d.file_path, d.file_name, d.file_type, d.file_size,
                           d.keywords, d.summary, d.category,
                           rank
                    FROM documents_fts f
                    JOIN documents d ON d.id = f.rowid
                    WHERE documents_fts MATCH ?
                    ORDER BY rank
                    LIMIT ?
                    """,
                    (query, top_k),
                ).fetchall()
            except sqlite3.OperationalError:
                # FTS5 non disponible, fallback LIKE
                like_query = f"%{query}%"
                rows = self._conn.execute(
                    """
                    SELECT file_path, file_name, file_type, file_size,
                           keywords, summary, category, 0
                    FROM documents
                    WHERE file_name LIKE ? OR keywords LIKE ?
                          OR summary LIKE ? OR content_preview LIKE ?
                    LIMIT ?
                    """,
                    (like_query, like_query, like_query, like_query, top_k),
                ).fetchall()

            for row in rows:
                # FTS5 rank est negatif (plus negatif = plus pertinent)
                rank = abs(float(row[7])) if row[7] else 1.0
                keyword_score = 1.0 / (1.0 + rank) if rank != 0 else 0.5

                results.append({
                    "file_path": row[0],
                    "file_name": row[1],
                    "file_type": row[2],
                    "file_size": row[3],
                    "keywords": json.loads(row[4]),
                    "summary": row[5],
                    "category": row[6],
                    "keyword_score": keyword_score,
                    "score": keyword_score,
                })

        return results

    async def search_by_keywords(
        self, keywords: List[str], top_k: int = 10
    ) -> List[Dict[str, Any]]:
        """Recherche par mots-cles uniquement."""
        query = " OR ".join(keywords)
        return await self._search_keywords(query, top_k)

    # ─────────────────────────────────────────────────────────────────────
    # Maintenance
    # ─────────────────────────────────────────────────────────────────────
    async def reindex_changed(self) -> int:
        """Re-indexe uniquement les fichiers modifies depuis la derniere indexation."""
        count = 0
        with self._lock:
            rows = self._conn.execute(
                "SELECT file_path, modified_at, content_hash FROM documents"
            ).fetchall()

        for file_path, old_mtime, old_hash in rows:
            p = Path(file_path)
            if not p.exists():
                continue
            try:
                current_mtime = p.stat().st_mtime
                if current_mtime > old_mtime:
                    current_hash = self._compute_hash(file_path)
                    if current_hash != old_hash:
                        doc = await self.index_file(file_path)
                        if doc is not None:
                            count += 1
            except OSError:
                continue

        if count > 0:
            logger.info(f"Re-indexe {count} fichiers modifies")
        return count

    async def remove_stale(self) -> int:
        """Supprime de l'index les fichiers qui n'existent plus."""
        count = 0
        with self._lock:
            rows = self._conn.execute(
                "SELECT id, file_path FROM documents"
            ).fetchall()

        stale_ids: List[int] = []
        for doc_id, file_path in rows:
            if not Path(file_path).exists():
                stale_ids.append(doc_id)
                count += 1

        if stale_ids:
            with self._lock:
                for doc_id in stale_ids:
                    try:
                        self._conn.execute(
                            "DELETE FROM documents_fts WHERE rowid = ?", (doc_id,)
                        )
                    except sqlite3.OperationalError:
                        pass
                    self._conn.execute(
                        "DELETE FROM documents WHERE id = ?", (doc_id,)
                    )
                self._conn.commit()
            logger.info(f"Supprime {count} fichiers obsoletes de l'index")

        return count

    async def get_index_stats(self) -> Dict[str, Any]:
        """Statistiques de l'index."""
        with self._lock:
            total = self._conn.execute(
                "SELECT COUNT(*) FROM documents"
            ).fetchone()
            by_type = self._conn.execute(
                "SELECT file_type, COUNT(*) FROM documents GROUP BY file_type ORDER BY COUNT(*) DESC"
            ).fetchall()
            by_category = self._conn.execute(
                "SELECT category, COUNT(*) FROM documents GROUP BY category ORDER BY COUNT(*) DESC"
            ).fetchall()
            total_size = self._conn.execute(
                "SELECT COALESCE(SUM(file_size), 0) FROM documents"
            ).fetchone()

        return {
            "total_files": total[0] if total else 0,
            "total_size_bytes": total_size[0] if total_size else 0,
            "by_type": {r[0]: r[1] for r in by_type},
            "by_category": {r[0]: r[1] for r in by_category},
            "faiss_vectors": self._faiss_index.ntotal,
        }

    def close(self) -> None:
        """Ferme les connexions."""
        try:
            self._save_faiss()
        except Exception:
            pass
        self._conn.close()
